#!/usr/bin/env python3
"""
California Business Licensing URL Database Generator
Creates comprehensive XLSX file from research data
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime

# Define all the researched URLs organized by department
# This is a comprehensive compilation of all research conducted

data = []

# ============================================================================
# CALIFORNIA SECRETARY OF STATE (SOS)
# ============================================================================
sos_urls = [
    ("California Secretary of State", "https://www.sos.ca.gov/", "Portal", "Main homepage for CA Secretary of State", "Primary entry point for all SOS services"),
    ("California Secretary of State", "https://www.sos.ca.gov/business-programs", "Portal", "Business Programs Division homepage", "Entry point for all business services"),
    ("California Secretary of State", "https://bizfileonline.sos.ca.gov/", "Application", "bizfile Online main application portal", "Primary online filing system for businesses; requires Okta account"),
    ("California Secretary of State", "https://bizfileonline.sos.ca.gov/search/business", "Search", "Business entity search", "Free search of 17M+ business records by name or entity number"),
    ("California Secretary of State", "https://www.sos.ca.gov/business-programs/business-entities/starting-business", "Guide", "Starting a business overview", "Step-by-step guide for new businesses"),
    ("California Secretary of State", "https://www.sos.ca.gov/business-programs/business-entities/statements", "Filing", "Statement of Information overview", "Filing requirements, deadlines, $250 late fee penalty"),
    ("California Secretary of State", "https://www.sos.ca.gov/business-programs/business-entities/name-reservations", "Filing", "Name reservation overview", "$10 fee, 60-day reservation period"),
    ("California Secretary of State", "https://www.sos.ca.gov/business-programs/business-entities/forms", "Forms", "All business entity forms", "Comprehensive forms library with samples"),
    ("California Secretary of State", "https://bpd.cdn.sos.ca.gov/pdf/be-fee-schedule-062018.pdf", "Fees", "Business entity fee schedule", "Complete fee list (PDF)"),
    ("California Secretary of State", "https://www.sos.ca.gov/business-programs/ucc", "Portal", "UCC main page", "Uniform Commercial Code filing services"),
    ("California Secretary of State", "https://www.sos.ca.gov/business-programs/ts", "Portal", "Trademarks & service marks", "State trademark registration ($70 per class)"),
    ("California Secretary of State", "https://calicodev.sos.ca.gov/", "API", "SOS API developer portal", "Programmatic access to business data; requires signup"),
    ("California Secretary of State", "https://www.sos.ca.gov/business-programs/business-entities/processing-dates", "Help", "Current processing dates", "Updated processing times by submission method"),
    ("California Secretary of State", "https://www.sos.ca.gov/business-programs/business-entities/faqs", "Help", "Business entities FAQs", "Comprehensive FAQ covering all entity types"),
    ("California Secretary of State", "https://www.sos.ca.gov/business-programs/business-entities/contact", "Contact", "Business Entities contact", "Sacramento: (916) 653-6814"),
]
data.extend(sos_urls)

# ============================================================================
# FRANCHISE TAX BOARD (FTB)
# ============================================================================
ftb_urls = [
    ("Franchise Tax Board", "https://www.ftb.ca.gov/", "Portal", "Main FTB homepage", "Primary entry point for CA tax services"),
    ("Franchise Tax Board", "https://www.ftb.ca.gov/myftb/index.asp", "Application", "MyFTB account login", "Online account for filing and payments"),
    ("Franchise Tax Board", "https://www.ftb.ca.gov/file/business/index.html", "Portal", "Business filing information", "Filing requirements for all business entity types"),
    ("Franchise Tax Board", "https://www.ftb.ca.gov/file/business/types/corporations/index.html", "Guide", "Corporation tax information", "C-Corp and S-Corp tax requirements"),
    ("Franchise Tax Board", "https://www.ftb.ca.gov/file/business/types/limited-liability-company/index.html", "Guide", "LLC tax information", "$800 minimum franchise tax; annual fee based on income"),
    ("Franchise Tax Board", "https://www.ftb.ca.gov/file/business/types/partnerships/index.html", "Guide", "Partnership tax information", "LP, LLP, and general partnership requirements"),
    ("Franchise Tax Board", "https://www.ftb.ca.gov/pay/index.html", "Application", "Web Pay system", "Online payment for taxes"),
    ("Franchise Tax Board", "https://www.ftb.ca.gov/forms/index.html", "Forms", "Forms and publications", "Searchable database of all FTB forms"),
    ("Franchise Tax Board", "https://www.ftb.ca.gov/file/business/credits/index.html", "Guide", "Business tax credits", "Available credits and deductions"),
    ("Franchise Tax Board", "https://www.ftb.ca.gov/help/contact/index.html", "Contact", "Contact information", "Phone: 1-800-852-5711"),
]
data.extend(ftb_urls)

# ============================================================================
# EMPLOYMENT DEVELOPMENT DEPARTMENT (EDD)
# ============================================================================
edd_urls = [
    ("Employment Development Department", "https://edd.ca.gov/", "Portal", "Main EDD homepage", "Primary entry point for employer services"),
    ("Employment Development Department", "https://edd.ca.gov/en/Payroll_Taxes/", "Portal", "Payroll taxes main page", "UI, SDI, ETT, PIT tax information"),
    ("Employment Development Department", "https://eddservices.edd.ca.gov/", "Application", "e-Services for Business portal", "Online employer account management"),
    ("Employment Development Department", "https://edd.ca.gov/en/Payroll_Taxes/Am_I_Required_to_Register_as_an_Employer/", "Guide", "Employer registration requirements", "Determine if registration is required"),
    ("Employment Development Department", "https://edd.ca.gov/en/Payroll_Taxes/Forms_and_Publications/", "Forms", "Forms and publications", "DE forms and employer guides"),
    ("Employment Development Department", "https://edd.ca.gov/en/Payroll_Taxes/Rates_and_Withholding/", "Guide", "Tax rates and withholding", "Current 2025-2026 tax rates"),
    ("Employment Development Department", "https://edd.ca.gov/en/Payroll_Taxes/File_and_Pay/", "Application", "File and pay", "Quarterly returns and deposits"),
    ("Employment Development Department", "https://edd.ca.gov/en/Payroll_Taxes/New_Hire_Reporting/", "Guide", "New hire reporting", "DE 34 requirements"),
    ("Employment Development Department", "https://edd.ca.gov/en/Payroll_Taxes/Independent_Contractor_versus_Employee/", "Guide", "Worker classification", "Employee vs independent contractor determination"),
    ("Employment Development Department", "https://edd.ca.gov/en/about_edd/contact_edd/", "Contact", "Contact information", "Taxpayer Assistance: 1-888-745-3886"),
]
data.extend(edd_urls)

# ============================================================================
# CDTFA (California Department of Tax and Fee Administration)
# ============================================================================
cdtfa_urls = [
    ("CDTFA", "https://cdtfa.ca.gov/", "Portal", "Main CDTFA homepage", "Sales tax, seller's permits, excise taxes"),
    ("CDTFA", "https://cdtfa.ca.gov/services/", "Application", "Online Services portal", "Filing, payments, registration"),
    ("CDTFA", "https://cdtfa.ca.gov/services/registration.htm", "Application", "Online registration", "Free seller's permit registration"),
    ("CDTFA", "https://cdtfa.ca.gov/taxes-and-fees/sutprograms.htm", "Guide", "Sales & Use Tax overview", "Base rate 7.25% plus district taxes"),
    ("CDTFA", "https://maps.cdtfa.ca.gov/", "Application", "Tax rate lookup by address", "Interactive map for address-based rates"),
    ("CDTFA", "https://services.cdtfa.ca.gov/webservices/verification.jsp", "Search", "Permit verification", "Verify permits, licenses, accounts"),
    ("CDTFA", "https://cdtfa.ca.gov/Industry/cannabis/", "Guide", "Cannabis tax guide", "Cannabis excise tax requirements"),
    ("CDTFA", "https://cdtfa.ca.gov/taxes-and-fees/cigarette-and-tobacco-products/", "Guide", "Cigarette & tobacco tax", "Licensing and tax requirements"),
    ("CDTFA", "https://cdtfa.ca.gov/formspubs/", "Forms", "Forms & publications", "All CDTFA forms and publications"),
    ("CDTFA", "https://cdtfa.ca.gov/contact.htm", "Contact", "Contact information", "Phone: 1-800-400-7115"),
]
data.extend(cdtfa_urls)

# ============================================================================
# DEPARTMENT OF CONSUMER AFFAIRS (DCA) - Main
# ============================================================================
dca_urls = [
    ("Department of Consumer Affairs", "https://www.dca.ca.gov/", "Portal", "Main DCA homepage", "Umbrella for 40+ licensing boards/bureaus"),
    ("Department of Consumer Affairs", "https://www.breeze.ca.gov/", "Application", "BreEZe licensing system", "Online licensing for many DCA entities"),
    ("Department of Consumer Affairs", "https://search.dca.ca.gov/", "Search", "DCA License Search", "Unified search across all DCA boards"),
    ("Department of Consumer Affairs", "https://www.dca.ca.gov/about_us/entities.shtml", "Directory", "Boards and bureaus directory", "Complete list of all DCA entities"),
    ("Department of Consumer Affairs", "https://www.dca.ca.gov/consumers/complaints/index.shtml", "Help", "File a complaint", "Consumer complaint process"),
    ("Department of Consumer Affairs", "https://www.dca.ca.gov/consumers/index.shtml", "Help", "Consumer information", "Consumer protection resources"),
]
data.extend(dca_urls)

# ============================================================================
# DCA BOARDS AND BUREAUS
# ============================================================================
dca_boards = [
    # Contractors State License Board
    ("Contractors State License Board (CSLB)", "https://www.cslb.ca.gov/", "Portal", "CSLB main homepage", "Contractor licensing for 43 classifications"),
    ("Contractors State License Board (CSLB)", "https://www.cslb.ca.gov/onlineservices/", "Application", "CSLB online services", "Applications, renewals, license lookup"),
    ("Contractors State License Board (CSLB)", "https://www.cslb.ca.gov/about_us/library/licensing_classifications/", "Guide", "License classifications", "All 43 contractor classifications"),
    ("Contractors State License Board (CSLB)", "https://www.cslb.ca.gov/consumers/license_check.aspx", "Search", "License lookup", "Verify contractor licenses"),

    # Board of Registered Nursing
    ("Board of Registered Nursing", "https://www.rn.ca.gov/", "Portal", "BRN main homepage", "RN and APRN licensing"),
    ("Board of Registered Nursing", "https://www.rn.ca.gov/online/verify.shtml", "Search", "License verification", "Verify nursing licenses"),
    ("Board of Registered Nursing", "https://www.rn.ca.gov/applicants/index.shtml", "Guide", "Applicant information", "License requirements"),

    # Medical Board of California
    ("Medical Board of California", "https://www.mbc.ca.gov/", "Portal", "MBC main homepage", "Physician and surgeon licensing"),
    ("Medical Board of California", "https://www.mbc.ca.gov/License-Verification/", "Search", "License verification", "Verify physician licenses"),
    ("Medical Board of California", "https://www.mbc.ca.gov/Licensing/", "Guide", "Licensing information", "Application requirements"),

    # California Board of Accountancy
    ("California Board of Accountancy", "https://www.dca.ca.gov/cba/", "Portal", "CBA main homepage", "CPA and firm licensing"),
    ("California Board of Accountancy", "https://www.dca.ca.gov/cba/applicants/", "Guide", "Applicant information", "CPA exam and license requirements"),

    # Bureau of Automotive Repair
    ("Bureau of Automotive Repair", "https://bar.ca.gov/", "Portal", "BAR main homepage", "Auto repair and smog station licensing"),
    ("Bureau of Automotive Repair", "https://bar.ca.gov/Industry/ard/How_to_Become_Licensed", "Guide", "Dealer licensing", "Auto repair dealer requirements"),
    ("Bureau of Automotive Repair", "https://bar.ca.gov/Industry/smog/How_to_Become_Licensed", "Guide", "Smog station licensing", "Smog check station requirements"),

    # Board of Pharmacy
    ("Board of Pharmacy", "https://www.pharmacy.ca.gov/", "Portal", "Pharmacy Board homepage", "Pharmacist and pharmacy licensing"),

    # Dental Board
    ("Dental Board of California", "https://www.dbc.ca.gov/", "Portal", "Dental Board homepage", "Dentist and dental auxiliary licensing"),

    # Board of Barbering and Cosmetology
    ("Board of Barbering and Cosmetology", "https://www.barbercosmo.ca.gov/", "Portal", "BBC main homepage", "Barber, cosmetologist, establishment licensing"),

    # Bureau of Security and Investigative Services
    ("Bureau of Security and Investigative Services", "https://www.bsis.ca.gov/", "Portal", "BSIS main homepage", "Security guard, PI, locksmith licensing"),

    # California Architects Board
    ("California Architects Board", "https://www.cab.ca.gov/", "Portal", "CAB main homepage", "Architect licensing"),

    # Board of Professional Engineers
    ("Board for Professional Engineers", "https://www.bpelsg.ca.gov/", "Portal", "BPELSG homepage", "Engineer, land surveyor, geologist licensing"),

    # Veterinary Medical Board
    ("Veterinary Medical Board", "https://www.vmb.ca.gov/", "Portal", "VMB main homepage", "Veterinarian and RVT licensing"),

    # Real Estate Appraisers Bureau
    ("Bureau of Real Estate Appraisers", "https://www.brea.ca.gov/", "Portal", "BREA main homepage", "Real estate appraiser licensing"),

    # Cemetery and Funeral Bureau
    ("Cemetery and Funeral Bureau", "https://www.cfb.ca.gov/", "Portal", "CFB main homepage", "Funeral director, embalmer, cemetery licensing"),

    # Physical Therapy Board
    ("Physical Therapy Board", "https://www.ptbc.ca.gov/", "Portal", "PTBC main homepage", "Physical therapist licensing"),

    # Board of Psychology
    ("Board of Psychology", "https://www.psychology.ca.gov/", "Portal", "BOP main homepage", "Psychologist licensing"),

    # Board of Behavioral Sciences
    ("Board of Behavioral Sciences", "https://www.bbs.ca.gov/", "Portal", "BBS main homepage", "LCSW, LMFT, LPCC, LEP licensing"),

    # Respiratory Care Board
    ("Respiratory Care Board", "https://www.rcb.ca.gov/", "Portal", "RCB main homepage", "Respiratory care practitioner licensing"),

    # Acupuncture Board
    ("Acupuncture Board", "https://www.acupuncture.ca.gov/", "Portal", "Acupuncture Board homepage", "Acupuncturist licensing"),

    # Board of Optometry
    ("Board of Optometry", "https://www.optometry.ca.gov/", "Portal", "Optometry Board homepage", "Optometrist licensing"),

    # Physician Assistant Board
    ("Physician Assistant Board", "https://www.pab.ca.gov/", "Portal", "PAB main homepage", "Physician assistant licensing"),

    # Podiatric Medical Board
    ("Podiatric Medical Board", "https://www.pmbc.ca.gov/", "Portal", "PMBC main homepage", "Podiatrist licensing"),

    # Structural Pest Control Board
    ("Structural Pest Control Board", "https://www.pestboard.ca.gov/", "Portal", "SPCB main homepage", "Pest control operator licensing"),

    # Court Reporters Board
    ("Court Reporters Board", "https://www.courtreportersboard.ca.gov/", "Portal", "CRB main homepage", "Court reporter licensing"),

    # Board of Chiropractic Examiners
    ("Board of Chiropractic Examiners", "https://www.chiro.ca.gov/", "Portal", "BCE main homepage", "Chiropractor licensing"),

    # Osteopathic Medical Board
    ("Osteopathic Medical Board", "https://www.ombc.ca.gov/", "Portal", "OMBC main homepage", "Osteopathic physician licensing"),

    # Dental Hygiene Board
    ("Dental Hygiene Board", "https://www.dhbc.ca.gov/", "Portal", "DHBC main homepage", "Dental hygienist licensing"),

    # Vocational Nursing Board
    ("Board of Vocational Nursing", "https://www.bvnpt.ca.gov/", "Portal", "BVNPT main homepage", "LVN and psychiatric technician licensing"),

    # Speech-Language Pathology Board
    ("Speech-Language Pathology Board", "https://www.slpab.ca.gov/", "Portal", "SLPAB main homepage", "Speech pathologist and audiologist licensing"),

    # Occupational Therapy Board
    ("Occupational Therapy Board", "https://www.bot.ca.gov/", "Portal", "CBOT main homepage", "Occupational therapist licensing"),

    # Professional Fiduciaries Bureau
    ("Professional Fiduciaries Bureau", "https://www.fiduciary.ca.gov/", "Portal", "PFB main homepage", "Professional fiduciary licensing"),

    # Bureau of Household Goods
    ("Bureau of Household Goods and Services", "https://bhgs.dca.ca.gov/", "Portal", "BHGS main homepage", "Mover, appliance repair licensing"),

    # Athletic Commission
    ("California State Athletic Commission", "https://www.dca.ca.gov/csac/", "Portal", "CSAC main homepage", "Boxing, MMA licensing"),

    # Private Postsecondary Education
    ("Bureau for Private Postsecondary Education", "https://www.bppe.ca.gov/", "Portal", "BPPE main homepage", "Private school approval"),
]
data.extend(dca_boards)

# ============================================================================
# DEPARTMENT OF ALCOHOLIC BEVERAGE CONTROL (ABC)
# ============================================================================
abc_urls = [
    ("Alcoholic Beverage Control", "https://www.abc.ca.gov/", "Portal", "Main ABC homepage", "Alcohol licensing for 70+ license types"),
    ("Alcoholic Beverage Control", "https://services.abc.ca.gov/", "Application", "ABC Online Services", "License applications, renewals, RBS"),
    ("Alcoholic Beverage Control", "https://www.abc.ca.gov/licensing/license-types/", "Guide", "License types", "All 70+ license type descriptions"),
    ("Alcoholic Beverage Control", "https://www.abc.ca.gov/licensing/license-lookup/", "Search", "License lookup", "Search by name, address, or license number"),
    ("Alcoholic Beverage Control", "https://www.abc.ca.gov/licensing/apply-for-a-new-license/", "Application", "Apply for new license", "Application process and forms"),
    ("Alcoholic Beverage Control", "https://www.abc.ca.gov/licensing/license-fees/", "Fees", "License fees", "Application and annual fee schedules"),
    ("Alcoholic Beverage Control", "https://www.abc.ca.gov/licensing/license-forms/", "Forms", "License forms", "All ABC forms library"),
    ("Alcoholic Beverage Control", "https://www.abc.ca.gov/education/rbs/", "Guide", "RBS Training Program", "Mandatory training for on-premises servers"),
    ("Alcoholic Beverage Control", "https://www.abc.ca.gov/contact/district-offices/", "Contact", "District offices", "All ABC field office locations"),
]
data.extend(abc_urls)

# ============================================================================
# DEPARTMENT OF CANNABIS CONTROL (DCC)
# ============================================================================
dcc_urls = [
    ("Department of Cannabis Control", "https://cannabis.ca.gov/", "Portal", "Main DCC homepage", "All commercial cannabis licensing"),
    ("Department of Cannabis Control", "https://cannabis.ca.gov/applicants/license-types/", "Guide", "License types", "Cultivation, manufacturing, retail, distribution, testing"),
    ("Department of Cannabis Control", "https://cannabis.ca.gov/applicants/how-to-apply/", "Guide", "How to apply", "Step-by-step application process"),
    ("Department of Cannabis Control", "https://search.cannabis.ca.gov/", "Search", "License search", "Search licensed cannabis businesses"),
    ("Department of Cannabis Control", "https://cannabis.ca.gov/applicants/application-license-fees/", "Fees", "Application and license fees", "Fee schedules by license type"),
    ("Department of Cannabis Control", "https://cannabis.ca.gov/licensees/track-and-trace/", "Guide", "METRC track and trace", "Inventory tracking requirements"),
    ("Department of Cannabis Control", "https://cannabis.ca.gov/applicants/ceqa-review-for-cannabis-businesses/", "Guide", "CEQA requirements", "Environmental review requirements"),
    ("Department of Cannabis Control", "https://cannabis.ca.gov/cannabis-laws/dcc-regulations/", "Regulations", "DCC regulations", "CCR Title 4 regulations"),
    ("Department of Cannabis Control", "https://cannabis.ca.gov/about-us/contact-us/", "Contact", "Contact information", "DCC contact and FAQ"),
]
data.extend(dcc_urls)

# ============================================================================
# DEPARTMENT OF REAL ESTATE (DRE)
# ============================================================================
dre_urls = [
    ("Department of Real Estate", "https://www.dre.ca.gov/", "Portal", "Main DRE homepage", "Real estate licensing"),
    ("Department of Real Estate", "https://secure.dre.ca.gov/", "Application", "eLicensing portal", "Online licensing transactions"),
    ("Department of Real Estate", "https://pplinfo2.dre.ca.gov/", "Search", "License lookup", "Public license information"),
    ("Department of Real Estate", "https://www.dre.ca.gov/examinees/requirementssales.html", "Guide", "Salesperson requirements", "Age 18+, 3 courses required"),
    ("Department of Real Estate", "https://www.dre.ca.gov/examinees/requirementsbroker.html", "Guide", "Broker requirements", "2 years experience, 8 courses"),
    ("Department of Real Estate", "https://www.dre.ca.gov/licensees/cerequirements.html", "Guide", "Continuing education", "45 hours for renewal"),
    ("Department of Real Estate", "https://www.dre.ca.gov/licensees/fees.html", "Fees", "Fee schedule", "Exam and license fees"),
    ("Department of Real Estate", "https://www.dre.ca.gov/forms/", "Forms", "DRE forms", "All RE forms library"),
    ("Department of Real Estate", "https://www.dre.ca.gov/contact.html", "Contact", "Contact information", "Phone: (877) 373-4542"),
]
data.extend(dre_urls)

# ============================================================================
# CALIFORNIA DEPARTMENT OF INSURANCE (CDI)
# ============================================================================
cdi_urls = [
    ("Department of Insurance", "https://www.insurance.ca.gov/", "Portal", "Main CDI homepage", "Insurance producer licensing"),
    ("Department of Insurance", "https://www.insurance.ca.gov/0200-industry/", "Portal", "Industry portal", "Licensing and compliance"),
    ("Department of Insurance", "https://cdicloud.insurance.ca.gov/cal", "Search", "License status inquiry", "Verify insurance licenses"),
    ("Department of Insurance", "https://www.insurance.ca.gov/0200-industry/0010-producer-online-services/", "Application", "Producer online services", "Applications, renewals, CE"),
    ("Department of Insurance", "https://www.insurance.ca.gov/0200-industry/0010-producer-online-services/0200-exam-info/", "Guide", "Exam information", "PSI exam scheduling"),
    ("Department of Insurance", "https://www.insurance.ca.gov/0200-industry/0050-renew-license/", "Guide", "License renewal", "Renewal requirements and CE"),
    ("Department of Insurance", "https://www.insurance.ca.gov/01-consumers/101-help/", "Help", "File a complaint", "Consumer complaint process"),
]
data.extend(cdi_urls)

# ============================================================================
# CALIFORNIA DEPARTMENT OF FOOD AND AGRICULTURE (CDFA)
# ============================================================================
cdfa_urls = [
    ("Department of Food and Agriculture", "https://www.cdfa.ca.gov/", "Portal", "Main CDFA homepage", "Agricultural permits and registrations"),
    ("Department of Food and Agriculture", "https://www.cdfa.ca.gov/ahfss/", "Portal", "Animal Health & Food Safety", "Livestock, dairy, meat inspection"),
    ("Department of Food and Agriculture", "https://www.cdfa.ca.gov/is/", "Portal", "Inspection Services", "Feed, fertilizer, organic programs"),
    ("Department of Food and Agriculture", "https://www.cdfa.ca.gov/plant/", "Portal", "Plant Health", "Nursery, seed, hemp programs"),
    ("Department of Food and Agriculture", "https://www.cdfa.ca.gov/dms/", "Portal", "Measurement Standards", "Weighmaster licensing"),
    ("Department of Food and Agriculture", "https://www.cdfa.ca.gov/is/ffldrs/", "Guide", "Feed and fertilizer licensing", "Feed, fertilizer registration"),
    ("Department of Food and Agriculture", "https://www.cdfa.ca.gov/plant/pe/nursery/", "Guide", "Nursery licensing", "Nursery stock registration"),
    ("Department of Food and Agriculture", "https://www.cdfa.ca.gov/is/organicprogram/", "Guide", "Organic program", "Organic certification"),
    ("Department of Food and Agriculture", "https://www.cdfa.ca.gov/plant/industrialhemp/", "Guide", "Industrial hemp", "Hemp cultivation registration"),
    ("Department of Food and Agriculture", "https://www.cdfa.ca.gov/exec/county/countymap/", "Directory", "County Ag Commissioners", "Local agriculture contacts"),
]
data.extend(cdfa_urls)

# ============================================================================
# DEPARTMENT OF INDUSTRIAL RELATIONS (DIR)
# ============================================================================
dir_urls = [
    ("Department of Industrial Relations", "https://www.dir.ca.gov/", "Portal", "Main DIR homepage", "Workplace safety, labor standards"),
    ("Department of Industrial Relations", "https://www.dir.ca.gov/dosh/", "Portal", "Cal/OSHA division", "Occupational safety permits"),
    ("Department of Industrial Relations", "https://www.dir.ca.gov/dosh/permits.html", "Guide", "Cal/OSHA permits", "Construction, elevator, boiler permits"),
    ("Department of Industrial Relations", "https://www.dir.ca.gov/dosh/esd/elevator-permits.html", "Guide", "Elevator permits", "Elevator installation permits"),
    ("Department of Industrial Relations", "https://www.dir.ca.gov/dosh/esd/pressure-vessel-permits.html", "Guide", "Pressure vessel permits", "Boiler and pressure vessel permits"),
    ("Department of Industrial Relations", "https://www.dir.ca.gov/dosh/amusement-rides.html", "Guide", "Amusement ride permits", "Ride installation and operation"),
    ("Department of Industrial Relations", "https://www.dir.ca.gov/dlse/", "Portal", "Labor Standards Enforcement", "Wage claims, labor contractor licensing"),
    ("Department of Industrial Relations", "https://www.dir.ca.gov/dlse/FarmLaborContractor.html", "Guide", "Farm labor contractor", "FLC licensing requirements"),
    ("Department of Industrial Relations", "https://www.dir.ca.gov/public-works/public-works.html", "Portal", "Public works", "Contractor registration, prevailing wage"),
    ("Department of Industrial Relations", "https://www.dir.ca.gov/public-works/contractor-registration.html", "Application", "Contractor registration", "Public works contractor registration"),
    ("Department of Industrial Relations", "https://www.dir.ca.gov/das/", "Portal", "Apprenticeship Standards", "Apprenticeship programs"),
    ("Department of Industrial Relations", "https://www.dir.ca.gov/dwc/", "Portal", "Workers' Compensation", "WC requirements"),
]
data.extend(dir_urls)

# ============================================================================
# CALIFORNIA AIR RESOURCES BOARD (CARB)
# ============================================================================
carb_urls = [
    ("California Air Resources Board", "https://ww2.arb.ca.gov/", "Portal", "Main CARB homepage", "Air quality regulations"),
    ("California Air Resources Board", "https://arber.arb.ca.gov/", "Application", "ARBER registration", "TRU and drayage truck registration"),
    ("California Air Resources Board", "https://ww2.arb.ca.gov/our-work/programs/truck-and-bus-regulation", "Guide", "Truck and Bus regulation", "Diesel truck emission standards"),
    ("California Air Resources Board", "https://ssl.arb.ca.gov/trucrs_reporting/login.php", "Application", "TRUCRS reporting", "Truck compliance reporting"),
    ("California Air Resources Board", "https://ww2.arb.ca.gov/our-work/programs/portable-equipment-registration-program-perp", "Guide", "PERP program", "Portable equipment registration"),
    ("California Air Resources Board", "https://cleantruckcheck.arb.ca.gov/", "Application", "Clean Truck Check", "Heavy-duty vehicle inspection"),
    ("California Air Resources Board", "https://ww2.arb.ca.gov/our-work/programs/cap-and-trade-program/about", "Guide", "Cap-and-Trade", "GHG allowance program"),
    ("California Air Resources Board", "https://ww2.arb.ca.gov/our-work/programs/low-carbon-fuel-standard", "Guide", "Low Carbon Fuel Standard", "LCFS program"),
    ("California Air Resources Board", "https://ww2.arb.ca.gov/california-air-districts", "Directory", "Air districts map", "35 local air district links"),
]
data.extend(carb_urls)

# ============================================================================
# STATE WATER RESOURCES CONTROL BOARD
# ============================================================================
water_urls = [
    ("State Water Resources Control Board", "https://www.waterboards.ca.gov/", "Portal", "Main Water Board homepage", "Water quality regulation"),
    ("State Water Resources Control Board", "https://smarts.waterboards.ca.gov/smarts/faces/SwSmartsLogin.xhtml", "Application", "SMARTS login", "Stormwater permit management"),
    ("State Water Resources Control Board", "https://www.waterboards.ca.gov/water_issues/programs/stormwater/", "Portal", "Stormwater program", "Industrial and construction permits"),
    ("State Water Resources Control Board", "https://www.waterboards.ca.gov/water_issues/programs/stormwater/igp.html", "Guide", "Industrial General Permit", "IGP requirements"),
    ("State Water Resources Control Board", "https://www.waterboards.ca.gov/water_issues/programs/stormwater/construction.html", "Guide", "Construction General Permit", "CGP requirements"),
    ("State Water Resources Control Board", "https://www.waterboards.ca.gov/ust/", "Portal", "Underground storage tanks", "UST program"),
    ("State Water Resources Control Board", "https://www.waterboards.ca.gov/drinking_water/", "Portal", "Drinking water", "Public water system permits"),
    # Regional Boards
    ("Regional Water Board - North Coast (R1)", "https://www.waterboards.ca.gov/northcoast/", "Portal", "Region 1 homepage", "North Coast region"),
    ("Regional Water Board - San Francisco Bay (R2)", "https://www.waterboards.ca.gov/sanfranciscobay/", "Portal", "Region 2 homepage", "SF Bay region"),
    ("Regional Water Board - Central Coast (R3)", "https://www.waterboards.ca.gov/centralcoast/", "Portal", "Region 3 homepage", "Central Coast region"),
    ("Regional Water Board - Los Angeles (R4)", "https://www.waterboards.ca.gov/losangeles/", "Portal", "Region 4 homepage", "Los Angeles region"),
    ("Regional Water Board - Central Valley (R5)", "https://www.waterboards.ca.gov/centralvalley/", "Portal", "Region 5 homepage", "Central Valley region"),
    ("Regional Water Board - Lahontan (R6)", "https://www.waterboards.ca.gov/lahontan/", "Portal", "Region 6 homepage", "Lahontan region"),
    ("Regional Water Board - Colorado River (R7)", "https://www.waterboards.ca.gov/coloradoriver/", "Portal", "Region 7 homepage", "Colorado River region"),
    ("Regional Water Board - Santa Ana (R8)", "https://www.waterboards.ca.gov/santaana/", "Portal", "Region 8 homepage", "Santa Ana region"),
    ("Regional Water Board - San Diego (R9)", "https://www.waterboards.ca.gov/sandiego/", "Portal", "Region 9 homepage", "San Diego region"),
]
data.extend(water_urls)

# ============================================================================
# DEPARTMENT OF TOXIC SUBSTANCES CONTROL (DTSC)
# ============================================================================
dtsc_urls = [
    ("Department of Toxic Substances Control", "https://dtsc.ca.gov/", "Portal", "Main DTSC homepage", "Hazardous waste regulation"),
    ("Department of Toxic Substances Control", "https://dtsc.ca.gov/permits/", "Portal", "Permits main page", "Hazardous waste facility permits"),
    ("Department of Toxic Substances Control", "https://dtsc.ca.gov/generators/", "Portal", "Generator information", "Hazardous waste generator requirements"),
    ("Department of Toxic Substances Control", "https://hwts.dtsc.ca.gov/", "Application", "Hazardous Waste Tracking System", "Manifest and ID tracking"),
    ("Department of Toxic Substances Control", "https://www.envirostor.dtsc.ca.gov/public/", "Search", "EnviroStor database", "Cleanup site tracking"),
    ("Department of Toxic Substances Control", "https://evq.dtsc.ca.gov/Register.aspx", "Application", "Electronic Verification Questionnaire", "Annual ID number verification"),
    ("Department of Toxic Substances Control", "https://dtsc.ca.gov/scp/safer-consumer-products-program-overview/", "Guide", "Safer Consumer Products", "Chemical restrictions program"),
    ("Department of Toxic Substances Control", "https://dtsc.ca.gov/site-mitigation/", "Portal", "Site mitigation", "Cleanup and brownfields"),
    ("Department of Toxic Substances Control", "https://dtsc.ca.gov/dtsc-laws-regulations/", "Regulations", "Laws and regulations", "Title 22 CCR Division 4.5"),
    ("Department of Toxic Substances Control", "https://dtsc.ca.gov/contact-information-3/", "Contact", "Contact information", "Phone: 800-618-6942"),
]
data.extend(dtsc_urls)

# ============================================================================
# LOCAL AIR QUALITY MANAGEMENT DISTRICTS (Major Districts)
# ============================================================================
air_districts = [
    ("South Coast AQMD", "https://www.aqmd.gov/", "Portal", "SCAQMD main homepage", "LA, Orange, Riverside, San Bernardino counties"),
    ("South Coast AQMD", "https://www.aqmd.gov/home/permits", "Portal", "Permits main page", "Authority to Construct, Permit to Operate"),
    ("Bay Area AQMD", "https://www.baaqmd.gov/", "Portal", "BAAQMD main homepage", "9 Bay Area counties"),
    ("Bay Area AQMD", "https://www.baaqmd.gov/permits", "Portal", "Permits main page", "Air quality permits"),
    ("San Joaquin Valley APCD", "https://www.valleyair.org/", "Portal", "SJVAPCD main homepage", "8 Central Valley counties"),
    ("San Joaquin Valley APCD", "https://www.valleyair.org/permits/", "Portal", "Permits main page", "Permit applications and forms"),
    ("Sacramento Metro AQMD", "https://www.airquality.org/", "Portal", "SMAQMD main homepage", "Sacramento County"),
    ("San Diego County APCD", "https://www.sdapcd.org/", "Portal", "SDAPCD main homepage", "San Diego County"),
    ("Ventura County APCD", "https://www.vcapcd.org/", "Portal", "VCAPCD main homepage", "Ventura County"),
    ("Santa Barbara County APCD", "https://www.ourair.org/", "Portal", "SBCAPCD main homepage", "Santa Barbara County"),
    ("Monterey Bay Air Resources District", "https://www.mbard.org/", "Portal", "MBARD main homepage", "Monterey, Santa Cruz, San Benito counties"),
    ("Imperial County APCD", "https://www.icapcd.org/", "Portal", "ICAPCD main homepage", "Imperial County"),
    ("Mojave Desert AQMD", "https://www.mdaqmd.ca.gov/", "Portal", "MDAQMD main homepage", "High desert region"),
    ("Antelope Valley AQMD", "https://www.avaqmd.ca.gov/", "Portal", "AVAQMD main homepage", "Lancaster/Palmdale area"),
    ("Great Basin Unified APCD", "https://www.gbuapcd.org/", "Portal", "GBUAPCD main homepage", "Alpine, Mono, Inyo counties"),
    ("Placer County APCD", "https://www.placer.ca.gov/2137/Air-Pollution-Control-District", "Portal", "Placer APCD homepage", "Placer County"),
    ("El Dorado County AQMD", "https://www.edcgov.us/Government/aqmd", "Portal", "El Dorado AQMD homepage", "El Dorado County"),
    ("Yolo-Solano AQMD", "https://www.ysaqmd.org/", "Portal", "YSAQMD main homepage", "Yolo and Solano counties"),
    ("Butte County AQMD", "https://bcaqmd.org/", "Portal", "BCAQMD main homepage", "Butte County"),
    ("Shasta County AQMD", "https://www.co.shasta.ca.us/index/aq_index", "Portal", "Shasta AQMD homepage", "Shasta County"),
    ("Tehama County APCD", "https://www.tehamacounty.ca.gov/government/departments/environmental-health/air-pollution-control-district", "Portal", "Tehama APCD homepage", "Tehama County"),
    ("Mendocino County AQMD", "https://www.mendocinoapcd.org/", "Portal", "Mendocino AQMD homepage", "Mendocino County"),
    ("Lake County AQMD", "https://lakeaqmd.org/", "Portal", "Lake AQMD homepage", "Lake County"),
    ("North Coast Unified AQMD", "https://ncuaqmd.org/", "Portal", "NCUAQMD main homepage", "Humboldt, Del Norte, Trinity counties"),
    ("San Luis Obispo County APCD", "https://www.slocleanair.org/", "Portal", "SLO APCD main homepage", "San Luis Obispo County"),
    ("Kern County (Eastern Kern APCD)", "https://www.kernair.org/", "Portal", "Eastern Kern APCD homepage", "Eastern Kern County"),
    ("Feather River AQMD", "https://www.fraqmd.org/", "Portal", "FRAQMD main homepage", "Yuba and Sutter counties"),
]
data.extend(air_districts)

# ============================================================================
# GO-BIZ (Governor's Office of Business and Economic Development)
# ============================================================================
gobiz_urls = [
    ("GO-Biz", "https://business.ca.gov/", "Portal", "Main GO-Biz homepage", "Business attraction and assistance"),
    ("GO-Biz", "https://www.calgold.ca.gov/", "Application", "CalGOLD permit finder", "Searchable permit database"),
    ("GO-Biz", "https://business.ca.gov/advantages/permit-and-regulatory-assistance/", "Guide", "Permit assistance", "Permit streamlining services"),
    ("GO-Biz", "https://business.ca.gov/california-competes-tax-credit/", "Guide", "California Competes Tax Credit", "$180M+ annually for business expansion"),
    ("GO-Biz", "https://business.ca.gov/advantages/incentives-grants-and-financing/", "Portal", "Incentives and financing", "State funding programs"),
    ("GO-Biz", "https://business.ca.gov/resources/international-affairs-and-trade/", "Portal", "International trade", "Export and FDI assistance"),
    ("GO-Biz", "https://business.ca.gov/about/contact-us/", "Contact", "Contact information", "Phone: 1-877-345-4633"),
]
data.extend(gobiz_urls)

# ============================================================================
# CalOSBA (California Office of the Small Business Advocate)
# ============================================================================
calosba_urls = [
    ("CalOSBA", "https://calosba.ca.gov/", "Portal", "Main CalOSBA homepage", "Small business assistance"),
    ("CalOSBA", "https://calosba.ca.gov/for-small-businesses-and-non-profits/", "Portal", "Small business resources", "Programs and support"),
    ("CalOSBA", "https://calosba.ca.gov/for-small-businesses-and-non-profits/small-business-centers/", "Directory", "Small business centers", "Network of 1,000+ advisors"),
    ("CalOSBA", "https://calosba.ca.gov/funding-grants-incentives/", "Portal", "Funding programs", "Grants and financing"),
    ("CalOSBA", "https://calosba.ca.gov/for-small-businesses-and-non-profits/permits-licenses-regulation/", "Guide", "Permits and licensing guide", "Regulatory navigation"),
    ("CalOSBA", "https://calosba.ca.gov/for-small-businesses-and-non-profits/set-up-your-business-in-california/", "Guide", "Starting a business", "Setup guide"),
    ("CalOSBA", "https://outsmartdisaster.calosba.ca.gov/", "Portal", "Outsmart Disaster", "Business resilience program"),
    ("CalOSBA", "https://calosba.ca.gov/about/contact-us/", "Contact", "Contact information", "Small business support"),
]
data.extend(calosba_urls)

# Create DataFrame
df = pd.DataFrame(data, columns=['Department', 'URL', 'Category', 'Description', 'AI_Agent_Notes'])

# Create Excel workbook with formatting
wb = Workbook()
ws = wb.active
ws.title = "CA Business Licensing URLs"

# Define styles
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
alt_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
wrap_alignment = Alignment(wrap_text=True, vertical='top')

# Write headers
headers = ['Department', 'URL', 'Category', 'Description', 'AI Agent Notes']
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center', vertical='center')

# Write data
for row_idx, row_data in enumerate(data, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = wrap_alignment
        if row_idx % 2 == 0:
            cell.fill = alt_fill

# Set column widths
ws.column_dimensions['A'].width = 35
ws.column_dimensions['B'].width = 65
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 50
ws.column_dimensions['E'].width = 60

# Freeze top row
ws.freeze_panes = 'A2'

# Add autofilter
ws.auto_filter.ref = ws.dimensions

# Create summary sheet
ws2 = wb.create_sheet("Summary by Department")

# Get unique departments and counts
dept_counts = df['Department'].value_counts().sort_index()

# Write summary
ws2.cell(row=1, column=1, value="Department").font = header_font
ws2.cell(row=1, column=1).fill = header_fill
ws2.cell(row=1, column=2, value="URL Count").font = header_font
ws2.cell(row=1, column=2).fill = header_fill

for row_idx, (dept, count) in enumerate(dept_counts.items(), 2):
    ws2.cell(row=row_idx, column=1, value=dept)
    ws2.cell(row=row_idx, column=2, value=count)

ws2.cell(row=len(dept_counts)+3, column=1, value="Total URLs")
ws2.cell(row=len(dept_counts)+3, column=2, value=len(data))
ws2.cell(row=len(dept_counts)+3, column=1).font = Font(bold=True)
ws2.cell(row=len(dept_counts)+3, column=2).font = Font(bold=True)

ws2.column_dimensions['A'].width = 45
ws2.column_dimensions['B'].width = 15

# Create category sheet
ws3 = wb.create_sheet("Summary by Category")
cat_counts = df['Category'].value_counts().sort_index()

ws3.cell(row=1, column=1, value="Category").font = header_font
ws3.cell(row=1, column=1).fill = header_fill
ws3.cell(row=1, column=2, value="URL Count").font = header_font
ws3.cell(row=1, column=2).fill = header_fill

for row_idx, (cat, count) in enumerate(cat_counts.items(), 2):
    ws3.cell(row=row_idx, column=1, value=cat)
    ws3.cell(row=row_idx, column=2, value=count)

ws3.column_dimensions['A'].width = 25
ws3.column_dimensions['B'].width = 15

# Add metadata sheet
ws4 = wb.create_sheet("About")
metadata = [
    ("Title", "California Business Licensing URL Database"),
    ("Version", "1.0"),
    ("Date Created", datetime.now().strftime("%Y-%m-%d")),
    ("Purpose", "Comprehensive URL reference for AI agents assisting with California business licensing"),
    ("Source", "California State Government Websites (.ca.gov)"),
    ("Total URLs", str(len(data))),
    ("Departments Covered", str(len(dept_counts))),
    ("Categories", str(len(cat_counts))),
    ("", ""),
    ("Notes", "This database is intended as a reference for AI agents to provide accurate URLs to users seeking California business licensing information."),
    ("", "URLs should be verified before use as government websites may change."),
    ("", "For the most current information, always refer to the official California state websites."),
]

for row_idx, (label, value) in enumerate(metadata, 1):
    ws4.cell(row=row_idx, column=1, value=label).font = Font(bold=True) if label else Font()
    ws4.cell(row=row_idx, column=2, value=value)

ws4.column_dimensions['A'].width = 25
ws4.column_dimensions['B'].width = 80

# Save workbook
output_path = os.path.join(os.path.expanduser("~"), "Documents/GitHub/gov-ai-dev/bizbot/BizAssessment/CA_Business_Licensing_URLs.xlsx")
wb.save(output_path)
print(f"Excel file created successfully: {output_path}")
print(f"Total URLs: {len(data)}")
print(f"Departments: {len(dept_counts)}")
print(f"Categories: {len(cat_counts)}")
